#!/usr/bin/env python3
"""
setup_grado.py — Inicialización de base de datos para el Gestor de Horarios

Lee config.json y un fichero CSV/Excel con las asignaturas del grado y genera
una base de datos SQLite lista para usar con servidor_horarios.py.

Uso (estructura plana, carpeta raíz):
  python3 setup_grado.py                              → genera plantilla CSV
  python3 setup_grado.py asignaturas.csv             → crea BD (config.json en raíz)
  python3 setup_grado.py asignaturas.csv --force     → sobreescribe sin confirmar

Uso (estructura por grados):
  python3 setup_grado.py grados/GIDI                 → genera plantilla en esa carpeta
  python3 setup_grado.py grados/GIDI asignaturas.csv → crea BD usando grados/GIDI/config.json
  python3 setup_grado.py grados/GIDI asignaturas.csv --force
"""

import csv
import json
import os
import sqlite3
import sys
from datetime import date, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config.json"

DIAS_SEMANA = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES']
MESES = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
         'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']


# ─── CONFIG ──────────────────────────────────────────────────────────────────

def load_config():
    if not CONFIG_PATH.exists():
        print(f"ERROR: No se encuentra {CONFIG_PATH}")
        sys.exit(1)
    with open(CONFIG_PATH, encoding='utf-8') as f:
        cfg = json.load(f)
    # Validaciones mínimas
    for key in ('degree', 'server', 'degree_structure', 'calendario'):
        if key not in cfg:
            print(f"ERROR: Falta la sección '{key}' en config.json")
            sys.exit(1)
    return cfg


# ─── PLANTILLA CSV ───────────────────────────────────────────────────────────

def generar_plantilla(cfg):
    """Genera un CSV de ejemplo con las columnas necesarias (en la carpeta raíz)."""
    degree = cfg['degree']['acronym']
    path = SCRIPT_DIR / f"asignaturas_{degree}.csv"
    ds = cfg['degree_structure']
    num_cursos = ds.get('num_cursos', 4)

    rows = []
    # Una asignatura de ejemplo por cuatrimestre y curso
    for curso in range(1, num_cursos + 1):
        for cuat in ['1C', '2C']:
            rows.append({
                'codigo':       f'{curso}{cuat}001',
                'nombre':       f'Asignatura ejemplo {curso}º {cuat}',
                'curso':        curso,
                'cuatrimestre': cuat,
                'creditos':     6,
                'af1':          45,
                'af2':          15,
                'af4':          0,
            })

    _escribir_plantilla_csv(path, rows, cfg['degree']['acronym'])


def generar_plantilla_en(cfg, output_dir):
    """Genera un CSV de ejemplo en la carpeta indicada (uso con estructura de grados)."""
    degree = cfg['degree']['acronym']
    path = Path(output_dir) / f"asignaturas_{degree}.csv"
    ds = cfg['degree_structure']
    num_cursos = ds.get('num_cursos', 4)

    rows = []
    for curso in range(1, num_cursos + 1):
        for cuat in ['1C', '2C']:
            rows.append({
                'codigo':       f'{curso}{cuat}001',
                'nombre':       f'Asignatura ejemplo {curso}º {cuat}',
                'curso':        curso,
                'cuatrimestre': cuat,
                'creditos':     6,
                'af1':          45,
                'af2':          15,
                'af4':          0,
            })

    _escribir_plantilla_csv(path, rows, degree)


def _escribir_plantilla_csv(path, rows, acronym):
    """Escribe el CSV de plantilla en la ruta indicada e imprime instrucciones."""
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['codigo','nombre','curso','cuatrimestre','creditos','af1','af2','af4'])
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ Plantilla generada: {path}")
    print()
    print("  Columnas:")
    print("    codigo       — código único de la asignatura (ej: 101001)")
    print("    nombre       — nombre completo de la asignatura")
    print("    curso        — número de curso (1, 2, 3, 4...)")
    print("    cuatrimestre — 1C o 2C")
    print("    creditos     — créditos ECTS (puede ser decimal: 4.5)")
    print("    af1          — horas de teoría")
    print("    af2          — horas de laboratorio")
    print("    af4          — horas de informática/aula informática")
    print()
    print(f"Edita '{path.name}' con las asignaturas reales y ejecuta:")
    print(f"  python3 setup_grado.py {path}")


# ─── LECTURA DE ASIGNATURAS ──────────────────────────────────────────────────

def leer_asignaturas(path_str):
    """Lee CSV o Excel. Devuelve lista de dicts con las columnas normalizadas."""
    path = Path(path_str)
    if not path.exists():
        print(f"ERROR: No se encuentra el fichero '{path}'")
        sys.exit(1)

    ext = path.suffix.lower()

    if ext == '.csv':
        rows = _leer_csv(path)
    elif ext in ('.xlsx', '.xls'):
        rows = _leer_excel(path)
    else:
        print(f"ERROR: Formato no soportado '{ext}'. Usa .csv, .xlsx o .xls")
        sys.exit(1)

    return _normalizar(rows, path)


def _leer_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return [row for row in reader]


def _leer_excel(path):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: Se necesita openpyxl para leer Excel.")
        print("  Instala con: pip install openpyxl --break-system-packages")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
    return rows


def _normalizar(rows, path):
    """Normaliza y valida las columnas del fichero."""
    # Mapeo de posibles nombres de columna a nombres internos
    alias = {
        'codigo':       ['codigo', 'código', 'code', 'cod'],
        'nombre':       ['nombre', 'name', 'asignatura', 'subject'],
        'curso':        ['curso', 'year', 'course', 'año'],
        'cuatrimestre': ['cuatrimestre', 'cuat', 'semester', 'semestre'],
        'creditos':     ['creditos', 'créditos', 'ects', 'credits'],
        'af1':          ['af1', 'teoria', 'teoría', 'theory', 'horas_teoria', 'h_teoria'],
        'af2':          ['af2', 'laboratorio', 'lab', 'laboratory', 'horas_lab', 'h_lab'],
        'af4':          ['af4', 'informatica', 'informática', 'info', 'aula_info', 'horas_info', 'h_info'],
        'af5':          ['af5', 'eval_continua', 'evaluacion_continua', 'h_af5'],
        'af6':          ['af6', 'eval_final', 'evaluacion_final', 'h_af6'],
    }

    if not rows:
        print(f"ERROR: El fichero '{path.name}' está vacío.")
        sys.exit(1)

    # Detectar columnas disponibles (case-insensitive)
    headers_lower = {k.lower().strip(): k for k in rows[0].keys()}
    col_map = {}
    for campo, posibles in alias.items():
        for p in posibles:
            if p in headers_lower:
                col_map[campo] = headers_lower[p]
                break

    if 'nombre' not in col_map:
        print(f"ERROR: No se encontró la columna 'nombre' en '{path.name}'")
        print(f"  Columnas encontradas: {list(rows[0].keys())}")
        sys.exit(1)
    if 'curso' not in col_map or 'cuatrimestre' not in col_map:
        print(f"ERROR: Faltan columnas 'curso' y/o 'cuatrimestre' en '{path.name}'")
        sys.exit(1)

    result = []
    for i, row in enumerate(rows, 2):
        nombre = str(row.get(col_map.get('nombre', ''), '')).strip()
        if not nombre:
            continue  # Fila vacía
        try:
            curso = int(str(row.get(col_map.get('curso', ''), 1)).strip())
        except ValueError:
            print(f"  Aviso fila {i}: 'curso' no es un número, se ignora")
            continue
        cuat_raw = str(row.get(col_map.get('cuatrimestre', ''), '1C')).strip().upper()
        cuat = '1C' if '1' in cuat_raw else '2C'

        def _num(campo, default=0):
            try: return int(float(str(row.get(col_map.get(campo, ''), default) or default)))
            except: return default
        def _float(campo, default=0.0):
            try: return float(str(row.get(col_map.get(campo, ''), default) or default))
            except: return default

        # Código: si no hay columna, generar automáticamente
        if 'codigo' in col_map:
            codigo = str(row.get(col_map['codigo'], '')).strip()
        else:
            codigo = ''
        if not codigo:
            codigo = f"{curso}{cuat[0]}{len(result)+1:03d}"

        result.append({
            'codigo':       codigo,
            'nombre':       nombre,
            'curso':        curso,
            'cuatrimestre': cuat,
            'creditos':     _float('creditos', 6.0),
            'af1':          _num('af1', 0),
            'af2':          _num('af2', 0),
            'af4':          _num('af4', 0),
            'af5':          _num('af5', 0),
            'af6':          _num('af6', 0),
        })

    print(f"  📋 {len(result)} asignaturas leídas de '{path.name}'")
    return result


# ─── GENERACIÓN DE SEMANAS Y FECHAS ─────────────────────────────────────────

def build_date_map(cal_cfg, num_semanas):
    """
    Genera el mapa fecha→semana para un cuatrimestre.
    Devuelve:
      semanas_info: lista de {numero, descripcion, lunes_iso}
      date_map:     {iso_date: {numero, dia}} para días lectivos
      festivos_set: set de fechas festivas/no-lectivas
    """
    inicio = date.fromisoformat(cal_cfg['inicio'])
    fin    = date.fromisoformat(cal_cfg['fin'])

    # Festivos y no-lectivos
    festivos_set = set()
    for f in cal_cfg.get('festivos', []):
        if isinstance(f, dict):
            festivos_set.add(f['fecha'])
        elif isinstance(f, str):
            festivos_set.add(f)

    # Vacaciones: conjunto de fechas
    vac_set = set()
    for v in cal_cfg.get('vacaciones', []):
        if isinstance(v, dict):
            ini_v = date.fromisoformat(v['inicio'])
            fin_v = date.fromisoformat(v['fin'])
        elif isinstance(v, str) and '/' in v:
            parts = v.split('/')
            ini_v = date.fromisoformat(parts[0])
            fin_v = date.fromisoformat(parts[1])
        else:
            continue
        cur = ini_v
        while cur <= fin_v:
            vac_set.add(cur.isoformat())
            cur += timedelta(days=1)

    # Avanzar hasta el primer lunes
    cur = inicio
    while cur.weekday() != 0:
        cur += timedelta(days=1)

    semanas_info = []
    date_map = {}
    sem_num = 0

    while sem_num < num_semanas and cur <= fin:
        lunes = cur
        viernes = cur + timedelta(days=4)

        # Si todos los días lectivos de la semana son vacaciones, saltarla
        dias_lectivos = [lunes + timedelta(days=i) for i in range(5)
                         if (lunes + timedelta(days=i)) <= fin]
        if dias_lectivos and all(d.isoformat() in vac_set for d in dias_lectivos):
            cur += timedelta(days=7)
            continue

        sem_num += 1
        fin_real = min(viernes, fin)

        # Descripción de la semana
        if lunes.month == fin_real.month:
            desc = f"SEMANA {sem_num}:  {lunes.day} {MESES[lunes.month]} A {fin_real.day} {MESES[fin_real.month]}"
        else:
            desc = f"SEMANA {sem_num}:  {lunes.day} {MESES[lunes.month]} A {fin_real.day} {MESES[fin_real.month]}"

        semanas_info.append({
            'numero':      sem_num,
            'descripcion': desc,
            'lunes_iso':   lunes.isoformat()
        })

        # Mapa de fechas para esta semana
        for i, dia_nombre in enumerate(DIAS_SEMANA):
            d = lunes + timedelta(days=i)
            if d > fin:
                break
            date_map[d.isoformat()] = {'numero': sem_num, 'dia': dia_nombre}

        cur += timedelta(days=7)

    return semanas_info, date_map, festivos_set | vac_set


# ─── CREACIÓN DE TABLAS ──────────────────────────────────────────────────────

def create_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS asignaturas (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo       TEXT NOT NULL,
            nombre       TEXT NOT NULL,
            curso        INTEGER DEFAULT NULL,
            cuatrimestre TEXT DEFAULT NULL
        );
        CREATE TABLE IF NOT EXISTS grupos (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            curso         INTEGER,
            cuatrimestre  TEXT,
            grupo         TEXT,
            aula          TEXT DEFAULT '',
            clave         TEXT
        );
        CREATE TABLE IF NOT EXISTS franjas (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            label  TEXT,
            orden  INTEGER
        );
        CREATE TABLE IF NOT EXISTS semanas (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            grupo_id    INTEGER REFERENCES grupos(id),
            numero      INTEGER,
            descripcion TEXT
        );
        CREATE TABLE IF NOT EXISTS clases (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            semana_id       INTEGER REFERENCES semanas(id),
            dia             TEXT,
            franja_id       INTEGER REFERENCES franjas(id),
            asignatura_id   INTEGER REFERENCES asignaturas(id),
            aula            TEXT DEFAULT '',
            subgrupo        TEXT DEFAULT '',
            observacion     TEXT DEFAULT '',
            es_no_lectivo   INTEGER DEFAULT 0,
            contenido       TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS fichas (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            asignatura_id INTEGER NOT NULL UNIQUE REFERENCES asignaturas(id) ON DELETE CASCADE,
            creditos      REAL DEFAULT 0,
            af1           INTEGER DEFAULT 0,
            af2           INTEGER DEFAULT 0,
            af4           INTEGER DEFAULT 0,
            af5           INTEGER DEFAULT 0,
            af6           INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS festivos_calendario (
            fecha       TEXT PRIMARY KEY,
            tipo        TEXT DEFAULT 'no_lectivo',
            descripcion TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS fichas_override (
            codigo    TEXT NOT NULL,
            grupo_key TEXT NOT NULL DEFAULT '',
            motivo    TEXT DEFAULT '',
            ts        TEXT DEFAULT '',
            PRIMARY KEY (codigo, grupo_key)
        );
        CREATE TABLE IF NOT EXISTS finales_excluidas (
            periodo     TEXT NOT NULL,
            curso       TEXT NOT NULL,
            asig_codigo TEXT NOT NULL,
            asig_nombre TEXT DEFAULT '',
            PRIMARY KEY (periodo, curso, asig_codigo)
        );
        CREATE TABLE IF NOT EXISTS examenes_finales (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha          TEXT NOT NULL,
            curso          TEXT NOT NULL,
            asig_nombre    TEXT DEFAULT '',
            asig_codigo    TEXT DEFAULT '',
            turno          TEXT DEFAULT 'mañana',
            observacion    TEXT DEFAULT '',
            auto_generated INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS asignaturas_destacadas (
            codigo    TEXT NOT NULL,
            grupo_num TEXT NOT NULL DEFAULT '',
            PRIMARY KEY (codigo, grupo_num)
        );
    """)
    conn.commit()


# ─── INSERCIÓN DE DATOS ──────────────────────────────────────────────────────

def insert_franjas(conn, franjas_cfg):
    for f in franjas_cfg:
        conn.execute(
            "INSERT INTO franjas (label, orden) VALUES (?,?)",
            (f['label'], f['orden'])
        )
    conn.commit()
    print(f"  ✅ {len(franjas_cfg)} franjas insertadas")


def insert_grupos(conn, ds):
    """Inserta los grupos según degree_structure.grupos_por_curso.
    Si degree_structure.aulas_por_curso existe, asigna aula a cada grupo
    (índice 0 → grupo 1, índice 1 → grupo 2, etc.).
    """
    grupos_cfg    = ds['grupos_por_curso']
    aulas_cfg     = ds.get('aulas_por_curso', {})   # {"1": ["PS5","PS6"], ...}
    count = 0
    for curso_str, cuats in sorted(grupos_cfg.items()):
        curso      = int(curso_str)
        aulas_list = aulas_cfg.get(curso_str, [])   # lista de aulas para este curso
        for cuat in ['1C', '2C']:
            n = cuats.get(cuat, 1)
            if n == 1:
                aula  = aulas_list[0] if aulas_list else ''
                clave = f"{curso}_{cuat}_grupo_unico"
                conn.execute(
                    "INSERT INTO grupos (curso, cuatrimestre, grupo, aula, clave) VALUES (?,?,?,?,?)",
                    (curso, cuat, 'unico', aula, clave)
                )
                count += 1
            else:
                for num in range(1, n + 1):
                    aula  = aulas_list[num - 1] if num - 1 < len(aulas_list) else ''
                    clave = f"{curso}_{cuat}_grupo_{num}"
                    conn.execute(
                        "INSERT INTO grupos (curso, cuatrimestre, grupo, aula, clave) VALUES (?,?,?,?,?)",
                        (curso, cuat, str(num), aula, clave)
                    )
                    count += 1
    conn.commit()
    print(f"  ✅ {count} grupos insertados")


def insert_semanas(conn, ds, calendario, fecha_maps):
    """
    Inserta semanas para cada grupo según su cuatrimestre.
    fecha_maps: {'1C': (semanas_info, date_map, fest_set), '2C': (...)}
    """
    num_semanas = ds.get('num_semanas', 16)
    grupos = conn.execute("SELECT id, cuatrimestre FROM grupos").fetchall()
    count = 0
    for g in grupos:
        cuat = g[1]
        semanas_info = fecha_maps[cuat][0]
        for s in semanas_info:
            conn.execute(
                "INSERT INTO semanas (grupo_id, numero, descripcion) VALUES (?,?,?)",
                (g[0], s['numero'], s['descripcion'])
            )
            count += 1
    conn.commit()
    print(f"  ✅ {count} semanas insertadas")


def insert_asignaturas_fichas(conn, asignaturas):
    """Inserta asignaturas y sus fichas docentes."""
    count = 0
    for a in asignaturas:
        cur = conn.execute(
            "INSERT INTO asignaturas (codigo, nombre, curso, cuatrimestre) VALUES (?,?,?,?)",
            (a['codigo'], a['nombre'], a.get('curso'), a.get('cuatrimestre'))
        )
        asig_id = cur.lastrowid
        conn.execute(
            "INSERT INTO fichas (asignatura_id, creditos, af1, af2, af4, af5, af6) VALUES (?,?,?,?,?,?,?)",
            (asig_id, a['creditos'], a['af1'], a['af2'], a['af4'], a.get('af5', 0), a.get('af6', 0))
        )
        count += 1
    conn.commit()
    print(f"  ✅ {count} asignaturas y fichas insertadas")


def insert_festivos_calendario(conn, calendario):
    """Puebla festivos_calendario desde el config."""
    count = 0
    for cuat in ['1C', '2C']:
        cfg = calendario.get(cuat, {})
        for f in cfg.get('festivos', []):
            if isinstance(f, dict):
                fecha, tipo, desc = f['fecha'], f.get('tipo', 'no_lectivo'), f.get('descripcion', '')
            elif isinstance(f, str):
                fecha, tipo, desc = f, 'festivo', ''
            else:
                continue
            conn.execute(
                "INSERT OR REPLACE INTO festivos_calendario (fecha, tipo, descripcion) VALUES (?,?,?)",
                (fecha, tipo, desc)
            )
            count += 1
        # Vacaciones
        for v in cfg.get('vacaciones', []):
            if isinstance(v, dict):
                ini_v = date.fromisoformat(v['inicio'])
                fin_v = date.fromisoformat(v['fin'])
                desc  = v.get('descripcion', 'Vacaciones')
            elif isinstance(v, str) and '/' in v:
                parts = v.split('/')
                ini_v = date.fromisoformat(parts[0])
                fin_v = date.fromisoformat(parts[1])
                desc  = 'Vacaciones'
            else:
                continue
            cur = ini_v
            while cur <= fin_v:
                conn.execute(
                    "INSERT OR REPLACE INTO festivos_calendario (fecha, tipo, descripcion) VALUES (?,?,?)",
                    (cur.isoformat(), 'no_lectivo', desc)
                )
                cur += timedelta(days=1)
                count += 1
    conn.commit()
    print(f"  ✅ {count} días en festivos_calendario")


def import_clases_desde_excel(conn, clases_importadas):
    """
    Inserta en la BD las clases importadas desde Excel de curso anterior.
    Mapea semana N → semana N del nuevo calendario.
    Respeta los días no-lectivos del nuevo año (no los sobreescribe).
    """
    if not clases_importadas:
        print("  ℹ️  No hay clases para importar.")
        return

    # Mapa franja_label → franja_id (coincidencia exacta y por prefijo "HH:MM")
    franjas_db = conn.execute("SELECT id, label FROM franjas ORDER BY orden").fetchall()
    franja_map = {}
    for f in franjas_db:
        label = f[1].strip()
        franja_map[label] = f[0]
        prefix = label[:5].rstrip()
        franja_map.setdefault(prefix, f[0])

    def get_franja_id(label):
        label = (label or '').strip()
        if label in franja_map:
            return franja_map[label]
        return franja_map.get(label[:5].rstrip())

    # Mapa (curso, cuatrimestre) → [grupo_id, …]
    grupos_db = conn.execute("SELECT id, curso, cuatrimestre FROM grupos").fetchall()
    grupos_map = {}
    for g in grupos_db:
        grupos_map.setdefault((g[1], g[2]), []).append(g[0])

    count_ok         = 0
    count_nolect     = 0
    count_nofranja   = 0
    count_nogrupo    = 0
    count_nosem      = 0

    for clase in clases_importadas:
        curso      = clase.get('curso')
        cuat       = clase.get('cuatrimestre')
        sem_num    = clase.get('semana')
        dia        = (clase.get('dia') or '').strip().upper()
        franja_lbl = clase.get('franja_label', '')
        codigo     = (clase.get('asig_codigo') or '').strip()
        nombre     = (clase.get('asig_nombre') or '').strip()
        tipo       = clase.get('tipo', '')
        subgrupo   = clase.get('subgrupo', '')
        aula_ov    = clase.get('aula_override', '')

        if not codigo or not sem_num or not dia:
            continue

        franja_id = get_franja_id(franja_lbl)
        if franja_id is None:
            count_nofranja += 1
            continue

        grupo_ids = grupos_map.get((curso, cuat), [])
        if not grupo_ids:
            count_nogrupo += 1
            continue

        # Buscar o crear asignatura
        asig_row = conn.execute(
            "SELECT id FROM asignaturas WHERE codigo=?", (codigo,)
        ).fetchone()
        if asig_row:
            asig_id = asig_row[0]
        else:
            cur = conn.execute(
                "INSERT INTO asignaturas (codigo, nombre, curso, cuatrimestre) VALUES (?,?,?,?)",
                (codigo, nombre, curso, cuat)
            )
            asig_id = cur.lastrowid
            conn.execute(
                "INSERT OR IGNORE INTO fichas "
                "(asignatura_id, creditos, af1, af2, af4) VALUES (?,6,0,0,0)",
                (asig_id,)
            )

        # Determinar campo aula (LAB / INFO / aula especial / vacío = teoría)
        aula      = aula_ov if aula_ov else tipo
        contenido = f"[{codigo}] {nombre}"
        if tipo:
            contenido += f" | {tipo}"
        if subgrupo:
            contenido += f" | Subgrupos: {subgrupo}"

        for grupo_id in grupo_ids:
            sem_row = conn.execute(
                "SELECT id FROM semanas WHERE grupo_id=? AND numero=?",
                (grupo_id, sem_num)
            ).fetchone()
            if not sem_row:
                count_nosem += 1
                continue
            sem_id = sem_row[0]

            existing = conn.execute(
                "SELECT id, es_no_lectivo FROM clases "
                "WHERE semana_id=? AND dia=? AND franja_id=?",
                (sem_id, dia, franja_id)
            ).fetchone()

            if existing:
                if existing[1]:          # día no-lectivo en el nuevo año → no tocar
                    count_nolect += 1
                    continue
                conn.execute(
                    "UPDATE clases SET asignatura_id=?, aula=?, subgrupo=?, "
                    "observacion='', es_no_lectivo=0, contenido=? WHERE id=?",
                    (asig_id, aula, subgrupo, contenido, existing[0])
                )
            else:
                conn.execute(
                    "INSERT INTO clases "
                    "(semana_id, dia, franja_id, asignatura_id, aula, subgrupo, "
                    " observacion, es_no_lectivo, contenido) "
                    "VALUES (?,?,?,?,?,?,'',0,?)",
                    (sem_id, dia, franja_id, asig_id, aula, subgrupo, contenido)
                )
            count_ok += 1

    conn.commit()
    print(f"  ✅ {count_ok} clases importadas desde Excel")
    if count_nolect:
        print(f"  ℹ️  {count_nolect} clases omitidas (día no-lectivo en nuevo curso)")
    if count_nofranja:
        print(f"  ⚠️  {count_nofranja} clases sin franja horaria coincidente")
    if count_nogrupo + count_nosem:
        print(f"  ⚠️  {count_nogrupo + count_nosem} clases sin grupo/semana coincidente")


def apply_no_lectivos(conn, fecha_maps):
    """
    Crea entradas en clases con es_no_lectivo=1 para los días festivos/no-lectivos
    dentro del rango lectivo de cada cuatrimestre.
    """
    festivos = conn.execute("SELECT fecha, tipo, descripcion FROM festivos_calendario").fetchall()
    franjas  = conn.execute("SELECT id FROM franjas ORDER BY orden").fetchall()

    count = 0
    for fest in festivos:
        fecha_iso = fest[0]
        for cuat, (semanas_info, date_map, _) in fecha_maps.items():
            if fecha_iso not in date_map:
                continue
            info = date_map[fecha_iso]
            num_sem = info['numero']
            dia     = info['dia']
            desc    = fest[2] or fest[1]

            # Semanas de todos los grupos de este cuatrimestre con este número
            grupos = conn.execute(
                "SELECT g.id FROM grupos g WHERE g.cuatrimestre = ?", (cuat,)
            ).fetchall()
            for grupo in grupos:
                sem = conn.execute(
                    "SELECT id FROM semanas WHERE grupo_id=? AND numero=?",
                    (grupo[0], num_sem)
                ).fetchone()
                if not sem:
                    continue
                sem_id = sem[0]
                for franja in franjas:
                    # Upsert: actualiza si ya existe, inserta si no
                    existing = conn.execute(
                        "SELECT id FROM clases WHERE semana_id=? AND dia=? AND franja_id=?",
                        (sem_id, dia, franja[0])
                    ).fetchone()
                    if existing:
                        conn.execute(
                            "UPDATE clases SET es_no_lectivo=1, observacion=? WHERE id=?",
                            (desc, existing[0])
                        )
                    else:
                        conn.execute(
                            """INSERT INTO clases
                               (semana_id, dia, franja_id, asignatura_id, aula, subgrupo,
                                observacion, es_no_lectivo, contenido)
                               VALUES (?,?,?,NULL,'','',?,1,'')""",
                            (sem_id, dia, franja[0], desc)
                        )
                        count += 1
    conn.commit()
    print(f"  ✅ {count} clases no-lectivas creadas")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    global CONFIG_PATH
    force = '--force' in sys.argv
    args  = [a for a in sys.argv[1:] if not a.startswith('--')]

    print("=" * 55)
    print("  setup_grado.py — Gestor de Horarios Universitarios")
    print("=" * 55)

    # Detectar si el primer argumento es una carpeta de grado
    grado_dir = None
    csv_arg   = None
    if args:
        p = Path(args[0])
        if p.is_dir():
            grado_dir = p
            args = args[1:]   # el resto será el CSV (opcional)
        # Si no es directorio, es el CSV directamente (modo clásico)

    if csv_arg is None and args:
        csv_arg = args[0]

    # Apuntar CONFIG_PATH a la carpeta del grado si se indicó
    if grado_dir:
        CONFIG_PATH = grado_dir / "config.json"
        if not CONFIG_PATH.exists():
            print(f"ERROR: No se encuentra config.json en '{grado_dir}'")
            sys.exit(1)

    cfg = load_config()
    degree  = cfg['degree']
    ds      = cfg['degree_structure']
    cal     = cfg['calendario']
    db_name = cfg['server']['db_name']

    # La BD se guarda en la carpeta del grado si se indicó, si no en la raíz
    db_path = (grado_dir / db_name) if grado_dir else (SCRIPT_DIR / db_name)

    print(f"\n📚 Grado: {degree['name']} ({degree['acronym']})")
    print(f"📁 Carpeta: {grado_dir or SCRIPT_DIR}")
    print(f"🗄️  Base de datos: {db_name}")
    print(f"📅 Curso: {cfg['server']['curso_label']}")
    print()

    # Modo plantilla
    if not csv_arg:
        print("ℹ️  No se proporcionó fichero de asignaturas.")
        print("  Generando plantilla...\n")
        # Guardar plantilla en carpeta del grado si se indicó
        if grado_dir:
            generar_plantilla_en(cfg, grado_dir)
        else:
            generar_plantilla(cfg)
        return

    # Modo creación de BD
    csv_path = csv_arg

    if db_path.exists() and not force:
        print(f"⚠️  Ya existe '{db_path}'.")
        resp = input("  ¿Sobreescribir? (s/N): ").strip().lower()
        if resp != 's':
            print("  Cancelado. Usa --force para sobreescribir sin confirmar.")
            return
        db_path.unlink()
        print()

    if db_path.exists() and force:
        db_path.unlink()

    # Leer asignaturas
    print("📋 Leyendo asignaturas...")
    asignaturas = leer_asignaturas(csv_path)
    print()

    # Generar mapas de fechas para cada cuatrimestre
    print("📅 Calculando semanas...")
    num_semanas = ds.get('num_semanas', 16)
    fecha_maps = {}
    for cuat in ['1C', '2C']:
        if cuat not in cal:
            print(f"  Aviso: no hay calendario para {cuat} en config.json")
            fecha_maps[cuat] = ([], {}, set())
            continue
        semanas_info, date_map, fest_set = build_date_map(cal[cuat], num_semanas)
        fecha_maps[cuat] = (semanas_info, date_map, fest_set)
        sem_real = len(semanas_info)
        print(f"  {cuat}: {sem_real} semanas ({cal[cuat]['inicio']} → {cal[cuat]['fin']})")
    print()

    # Crear BD en /tmp para evitar errores de I/O en rutas de red (Dropbox, etc.)
    # y copiar al destino final al terminar
    import shutil
    tmp_path = Path('/tmp') / db_name
    if tmp_path.exists():
        tmp_path.unlink()

    print(f"🗄️  Creando '{db_name}' (trabajando en /tmp para evitar errores de red)...")
    conn = sqlite3.connect(str(tmp_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")

    create_tables(conn)
    insert_franjas(conn, ds['franjas'])
    insert_grupos(conn, ds)
    insert_semanas(conn, ds, cal, fecha_maps)
    insert_asignaturas_fichas(conn, asignaturas)
    insert_festivos_calendario(conn, cal)
    apply_no_lectivos(conn, fecha_maps)

    conn.execute("PRAGMA wal_checkpoint(FULL)")
    conn.close()

    # Copiar a la carpeta del proyecto
    shutil.copy2(str(tmp_path), str(db_path))
    tmp_path.unlink()

    # Resumen
    print()
    print("=" * 55)
    print(f"✅ Base de datos creada: {db_name}")
    print()
    print("  Para arrancar el servidor:")
    print(f'  python3 servidor_horarios.py')
    if db_name != 'horarios.db':
        print(f'  DB_PATH_OVERRIDE="{db_path}" python3 servidor_horarios.py')
    print()


if __name__ == '__main__':
    main()
