#!/usr/bin/env python3
"""
exportar_institucional.py — Exporta horarios al formato de la BD institucional UPCT.

Genera un Excel (.xlsx) con una fila por cada clase consolidada:
la misma (asignatura, tipo, día, franja, grupo, subgrupo) que se repite
en varias semanas aparece en UNA SOLA FILA con las semanas separadas por comas.

Columnas de salida (15):
  EVENTNAME, EVENTTYPE, DAYNUMBER, DAY, HOURBEGIN, HOUREND,
  TYPOLOGYNAME, ID_PLAN, MODULECODE, SUBGRUPO, STUDENTGROUPSECTION,
  TEACHERCODE, CLASSROOMS, WEEKS, ANNOTATIONS

Flujo de semanas:
  El script trabaja internamente sobre config/weeks.json (generado automáticamente
  desde weeks.xls si no existe). El JSON se guarda en config/ para poder
  editarse manualmente cuando sea necesario.
  Formato weeks.json: {"2026-09-07": 1, "2026-09-14": 2, ...}

Tipos de actividad:
  Se resuelven contra config/tipos_actividad.json. Para cada AF (AF1, AF2…)
  se toma el primer código que aparece en ese fichero. Si no hay coincidencia
  se asigna 'AD' por defecto.

Uso standalone:
  python3 tools/exportar_institucional.py \\
      --grado horarios/GIM \\
      --weeks weeks.xls            # (o weeks.json)
      [--output exportacion_institucional.xlsx]

Uso desde servidor:
  import exportar_institucional
  path = exportar_institucional.exportar(db_path, config_path, weeks_path, output_path)
"""

import os
import sys
import re
import json
import sqlite3
import argparse
from datetime import date, timedelta
from collections import defaultdict

# ── Dependencias opcionales ──────────────────────────────────────────────────
try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

SCRIPT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Constantes de mapeo ──────────────────────────────────────────────────────

DAY_NUMBER = {
    'LUNES': 1, 'MARTES': 2, 'MIÉRCOLES': 3, 'JUEVES': 4, 'VIERNES': 5
}
DAY_EN = {
    'LUNES': 'Monday', 'MARTES': 'Tuesday',
    'MIÉRCOLES': 'Wednesday', 'JUEVES': 'Thursday', 'VIERNES': 'Friday'
}
MESES_ES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
}
# Código por defecto cuando no hay coincidencia de AF
DEFAULT_TYPE_CODE = 'AD'

# Offset aplicado a los números de semana en la columna WEEKS del Excel institucional.
# La UPCT usa una numeración que comienza en 101 (semana 1 interna → 101 exportada).
WEEK_OFFSET = 100

HEADERS = [
    'EVENTNAME', 'EVENTTYPE', 'DAYNUMBER', 'DAY',
    'HOURBEGIN', 'HOUREND', 'TYPOLOGYNAME',
    'ID_PLAN', 'MODULECODE', 'SUBGRUPO',
    'STUDENTGROUPSECTION', 'TEACHERCODE',
    'CLASSROOMS', 'WEEKS', 'ANNOTATIONS'
]


# ═══════════════════════════════════════════════════════════════════════════════
# CARGA DE DATOS AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_weeks_xls(weeks_xls_path):
    """
    Lee weeks.xls y devuelve {date_str_lunes: semana_real (int)}.
    Uso interno: llamar load_weeks_map() en lugar de esta función.
    """
    if xlrd is None:
        raise ImportError("Se requiere xlrd: pip install xlrd --break-system-packages")
    wb = xlrd.open_workbook(weeks_xls_path)
    ws = wb.sheet_by_index(0)
    result = {}
    for row_idx in range(1, ws.nrows):
        row = ws.row_values(row_idx)
        if not row[0] or not row[2]:
            continue
        start_dt = xlrd.xldate_as_datetime(row[0], wb.datemode)
        week_num = int(row[2])
        result[start_dt.strftime('%Y-%m-%d')] = week_num
    return result


def _canonical_json_path():
    """Ruta canónica de weeks.json: siempre en config/ del proyecto."""
    return os.path.join(SCRIPT_DIR, 'config', 'weeks.json')


def load_weeks_map(weeks_path):
    """
    Carga el mapa de semanas reales desde weeks.json o weeks.xls.

    - Si weeks_path termina en .json → lo carga directamente.
    - Si termina en .xls/.xlsx → busca config/weeks.json en el proyecto.
      Si no existe, genera config/weeks.json desde el .xls y lo guarda allí.
    - El JSON tiene formato: {"2026-09-07": 1, "2026-09-14": 2, ...}
      y puede editarse manualmente para ajustar casos especiales.

    Devuelve {date_str: semana_real (int)}.
    """
    ext = os.path.splitext(weeks_path)[1].lower()

    if ext == '.json':
        if not os.path.exists(weeks_path):
            raise FileNotFoundError(f"No se encuentra {weeks_path}")
        with open(weeks_path, encoding='utf-8') as f:
            raw = json.load(f)
        return {k: int(v) for k, v in raw.items()}

    # Es un .xls/.xlsx → usar config/weeks.json como destino canónico
    json_path = _canonical_json_path()

    if not os.path.exists(json_path):
        # Generar config/weeks.json desde el .xls
        if not os.path.exists(weeks_path):
            raise FileNotFoundError(f"No se encuentra {weeks_path}")
        weeks_map = _parse_weeks_xls(weeks_path)
        ordered = dict(sorted(weeks_map.items()))
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(ordered, f, indent=2, ensure_ascii=False)
        print(f"  ✓ weeks.json generado en: {json_path}")
        return weeks_map

    # Ya existe config/weeks.json → cargarlo
    with open(json_path, encoding='utf-8') as f:
        raw = json.load(f)
    return {k: int(v) for k, v in raw.items()}


def load_tipos_actividad(tipos_json_path):
    """
    Carga config/tipos_actividad.json y construye el mapa AF → código institucional.

    Para cada AF, se toma el primer código que aparece en el fichero.
    Devuelve dict: {'AF1': 'AD', 'AF2': 'LAB', 'AF3': 'SEM', 'AF4': 'INF', ...}
    """
    if not os.path.exists(tipos_json_path):
        return {}
    with open(tipos_json_path, encoding='utf-8') as f:
        tipos = json.load(f)
    af_to_code = {}
    for entry in tipos:
        af = entry.get('af', '')
        code = entry.get('code', '')
        if af and code and af not in af_to_code:
            af_to_code[af] = code  # Primer código encontrado para cada AF
    return af_to_code


def load_classrooms(classrooms_json_path):
    """
    Carga classrooms.json y devuelve dos dicts:
      - by_code:  {'PS2': 'ETSII#PS2', ...}
      - by_name:  {'ETSII#PS2': 'ETSII#PS2', ...}
    """
    if not os.path.exists(classrooms_json_path):
        return {}, {}
    with open(classrooms_json_path, encoding='utf-8') as f:
        data = json.load(f)
    by_code = {}
    by_name = {}
    for entry in data:
        name = entry.get('ClassroomName', '')
        code = entry.get('ClassroomCode', '')
        if name:
            by_name[name] = name
        if code:
            by_code[code] = name or code
    return by_code, by_name


def load_config(config_path):
    """Carga config.json del grado."""
    cfg_file = config_path if config_path.endswith('.json') else os.path.join(config_path, 'config.json')
    with open(cfg_file, encoding='utf-8') as f:
        return json.load(f)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def parse_semana_start(descripcion, cuat_year):
    """
    Extrae la fecha lunes de la descripción de semana.
    Ej: "SEMANA 1:  7 SEPTIEMBRE A 11 SEPTIEMBRE" → date(2026, 9, 7)
    cuat_year: año del cuatrimestre (int), ej. 2026 para 1C, 2027 para 2C.
    """
    # Busca el primer "DD MES" después de ":"
    m = re.search(r':\s+(\d{1,2})\s+([A-ZÁÉÍÓÚ]+)', descripcion, re.IGNORECASE)
    if not m:
        return None
    day = int(m.group(1))
    month_str = m.group(2).upper().replace('É', 'E').replace('Á', 'A').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    # Normalización básica de acentos para el lookup
    month_str_norm = (descripcion[m.start(2):m.end(2)].upper()
                      .replace('É','E').replace('Á','A')
                      .replace('Í','I').replace('Ó','O').replace('Ú','U'))
    # Buscar en MESES_ES con y sin acentos
    month = None
    for k, v in MESES_ES.items():
        k_norm = k.replace('É','E').replace('Á','A').replace('Í','I').replace('Ó','O').replace('Ú','U')
        if k == m.group(2).upper() or k_norm == month_str_norm:
            month = v
            break
    if month is None:
        return None
    try:
        return date(cuat_year, month, day)
    except ValueError:
        return None


def get_typology(aula, activity_types, af_to_code):
    """
    Determina el código TYPOLOGYNAME a partir del campo aula de Janux.

    Regla principal: si el aula está vacía (no se especificó tipo en el horario)
    se asigna DEFAULT_TYPE_CODE ('AD') directamente.

    Para aulas con valor, se detecta el AF a través de activity_types (config.json)
    y se obtiene el código institucional de af_to_code (tipos_actividad.json).
    Si el código del AF es el mismo que el DEFAULT (ej. AF1 → 'AD'), se mantiene.
    Si no hay coincidencia en ningún mapa, se devuelve DEFAULT_TYPE_CODE.

    af_to_code: {'AF2': 'LAB', 'AF3': 'SEM', 'AF4': 'INF', ...}
    """
    if aula is None:
        aula = ''

    # Aula vacía → tipo no especificado → código por defecto
    if aula == '':
        return DEFAULT_TYPE_CODE

    # Aula con valor → buscar en activity_types
    for af, rules in activity_types.items():
        if not isinstance(rules, dict):
            continue  # Omite entradas no-dict como '_comment'
        if rules.get('fichas_only'):
            continue
        if aula in rules.get('aula_exact', []):
            return af_to_code.get(af, DEFAULT_TYPE_CODE)
        for prefix in rules.get('aula_startswith', []):
            if aula.startswith(prefix):
                return af_to_code.get(af, DEFAULT_TYPE_CODE)

    return DEFAULT_TYPE_CODE  # Sin coincidencia → Actividad Docente por defecto


def resolve_classroom(clase_aula, grupo_aula_default, by_code, by_name):
    """
    Devuelve el código institucional del aula.
    - Si clase_aula está vacía → usa grupo_aula_default
    - Busca en classrooms.json por nombre o código
    - Aplica transformación 'Aulario_PS#N' → 'PSN' como fallback
    """
    raw = clase_aula if clase_aula else (grupo_aula_default or '')

    # Lookup directo por nombre completo
    if raw in by_name:
        return raw

    # Lookup por código (e.g. 'PS2', 'LAB_INSTRU')
    if raw in by_code:
        return by_code[raw]

    # Transformación 'Aulario_PS#N' → 'PSN'
    m = re.match(r'Aulario[_\s]*PS#?(\d+)', raw, re.IGNORECASE)
    if m:
        code = f'PS{m.group(1)}'
        if code in by_code:
            return by_code[code]
        return f'ETSII#PS{m.group(1)}'

    # Transformación 'Aulario_PB#N' → 'PBN'
    m = re.match(r'Aulario[_\s]*PB#?(\d+)', raw, re.IGNORECASE)
    if m:
        code = f'PB{m.group(1)}'
        if code in by_code:
            return by_code[code]
        return f'ETSII#PB{m.group(1)}'

    return raw  # Sin transformación disponible


def eventname(nombre):
    """Genera el nombre del evento: 'Clase_' + nombre con espacios → '_'."""
    return 'Clase_' + nombre.replace(' ', '_')


def parse_franja(label):
    """
    Extrae inicio y fin de la etiqueta de franja.
    Ej: '9:00 - 10:50' → ('9:00', '10:50')
    """
    parts = label.split(' - ')
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return label.strip(), label.strip()


def section_code(acronym, curso, grupo, subg):
    """
    Genera el código de sección: GIM1_G1_4
    curso: int, grupo: str ('1','2',...), subg: str ('1','2',...)
    """
    return f'{acronym}{curso}_G{grupo}_{subg}'


# ═══════════════════════════════════════════════════════════════════════════════
# CONSULTA DB
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_data(db_path):
    """Consulta todos los datos necesarios del DB de Janux."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Grupos
    cur.execute("SELECT * FROM grupos ORDER BY curso, cuatrimestre, grupo")
    grupos = [dict(r) for r in cur.fetchall()]

    # Franjas
    cur.execute("SELECT id, label, orden FROM franjas ORDER BY orden")
    franjas = {r['id']: dict(r) for r in cur.fetchall()}

    # Por grupo: semanas, no-lectivos y clases
    data_by_grupo = {}

    for g in grupos:
        clave = g['clave']

        # Semanas del grupo
        cur.execute("""
            SELECT s.id, s.numero, s.descripcion
            FROM semanas s
            JOIN grupos gr ON s.grupo_id = gr.id
            WHERE gr.clave = ?
            ORDER BY s.numero
        """, (clave,))
        semanas = {r['numero']: dict(r) for r in cur.fetchall()}

        # Días no-lectivos: {(semana_num, dia)}
        cur.execute("""
            SELECT s.numero, c.dia
            FROM clases c
            JOIN semanas s ON c.semana_id = s.id
            JOIN grupos gr ON s.grupo_id = gr.id
            WHERE gr.clave = ? AND c.es_no_lectivo = 1
        """, (clave,))
        no_lectivos = set((r['numero'], r['dia']) for r in cur.fetchall())

        # Clases reales (no no-lectivos)
        # Nota: 'tipo' es la columna de tipo de actividad (LAB/INF/AD/EXP…);
        #        'aula' es el código de aula real (puede estar vacío → usar aula del grupo).
        # Comprobamos si existe la columna 'tipo' (introducida en sesión 11).
        cols_clases = {r['name'] for r in cur.execute('PRAGMA table_info(clases)').fetchall()}
        tipo_select = "c.tipo" if 'tipo' in cols_clases else "''"
        cur.execute(f"""
            SELECT c.dia, c.aula, c.subgrupo, c.observacion,
                   {tipo_select} AS tipo,
                   c.franja_id, f.label AS franja_label,
                   a.codigo, a.nombre,
                   s.numero AS semana_num
            FROM clases c
            JOIN semanas s ON c.semana_id = s.id
            JOIN grupos gr ON s.grupo_id = gr.id
            LEFT JOIN asignaturas a ON c.asignatura_id = a.id
            LEFT JOIN franjas f ON c.franja_id = f.id
            WHERE gr.clave = ?
              AND c.es_no_lectivo = 0
              AND a.codigo IS NOT NULL
            ORDER BY s.numero, c.dia, f.orden
        """, (clave,))
        clases = [dict(r) for r in cur.fetchall()]

        # Subgrupos numéricos distintos (excluye '' y 'todos')
        cur.execute("""
            SELECT DISTINCT c.subgrupo
            FROM clases c
            JOIN semanas s ON c.semana_id = s.id
            JOIN grupos gr ON s.grupo_id = gr.id
            WHERE gr.clave = ?
              AND c.es_no_lectivo = 0
              AND c.subgrupo != ''
              AND c.subgrupo != 'todos'
              AND c.subgrupo IS NOT NULL
        """, (clave,))
        subgrupos_numericos = sorted(
            [r['subgrupo'] for r in cur.fetchall()],
            key=lambda x: (int(x) if x.isdigit() else 999, x)
        )

        data_by_grupo[clave] = {
            'grupo': g,
            'semanas': semanas,
            'no_lectivos': no_lectivos,
            'clases': clases,
            'subgrupos_numericos': subgrupos_numericos,
        }

    conn.close()
    return grupos, data_by_grupo


# ═══════════════════════════════════════════════════════════════════════════════
# NÚCLEO: GENERACIÓN DE FILAS
# ═══════════════════════════════════════════════════════════════════════════════

def build_rows(grupos, data_by_grupo, config, weeks_map, by_code, by_name, af_to_code=None):
    """
    Genera todas las filas del Excel institucional.
    Devuelve lista de dicts con las 15 columnas.
    """
    acronym = config.get('degree', {}).get('acronym', 'GIM')
    activity_types = config.get('activity_types', {})
    calendario = config.get('calendario', {})

    # Año por cuatrimestre: extraído de config.calendario.<CUAT>.inicio
    cuat_year = {}
    for cuat, cal in calendario.items():
        if isinstance(cal, dict) and 'inicio' in cal:
            try:
                cuat_year[cuat] = int(cal['inicio'][:4])
            except (ValueError, TypeError):
                pass

    all_rows = []

    for g in grupos:
        clave = g['clave']
        gdata = data_by_grupo[clave]
        semanas = gdata['semanas']
        no_lectivos = gdata['no_lectivos']
        clases = gdata['clases']
        subgrupos_numericos = gdata['subgrupos_numericos']

        curso = g['curso']
        grupo_num = g['grupo']
        cuat = g['cuatrimestre']
        grupo_aula_default = g.get('aula', '')

        year = cuat_year.get(cuat)

        # ── Mapeo semana_interna → semana_real ──────────────────────────────
        semana_to_real = {}
        for sem_num, sem_info in semanas.items():
            if year is None:
                continue
            d = parse_semana_start(sem_info['descripcion'], year)
            if d is None:
                continue
            d_str = d.strftime('%Y-%m-%d')
            real = weeks_map.get(d_str)
            if real is not None:
                semana_to_real[sem_num] = real
            else:
                # Buscar la semana del lunes de esa fecha (por si hay desfase de días)
                # Retrocede al lunes de la semana
                lunes = d - timedelta(days=d.weekday())
                lunes_str = lunes.strftime('%Y-%m-%d')
                real2 = weeks_map.get(lunes_str)
                if real2 is not None:
                    semana_to_real[sem_num] = real2

        # ── Consolidar clases por clave ──────────────────────────────────────
        # Key: (nombre, codigo, typologyname, dia, franja_label, subgrupo_norm, classroom)
        # Value: list of real_week_nums
        consolidated = defaultdict(list)
        # También guardar metadata por key
        key_meta = {}

        for cls in clases:
            sem_num = cls['semana_num']
            dia = cls['dia']

            # Omitir si es un día no-lectivo
            if (sem_num, dia) in no_lectivos:
                continue

            real_week = semana_to_real.get(sem_num)
            if real_week is None:
                continue  # semana sin mapeo real (ej. vacaciones)

            aula = cls['aula'] or ''
            tipo = cls['tipo'] or ''          # tipo de actividad: LAB / INF / AD / EXP / …
            subgrupo = cls['subgrupo'] or ''
            nombre = cls['nombre'] or ''
            codigo = cls['codigo'] or ''
            franja_label = cls['franja_label'] or ''
            observacion = cls['observacion'] or ''

            # TYPOLOGYNAME: usar columna 'tipo' si está rellena;
            # si vacía, derivar de 'aula' (compatibilidad con BDs sin columna tipo).
            if tipo:
                typology = tipo
            else:
                typology = get_typology(aula, activity_types, af_to_code or {})

            classroom = resolve_classroom(aula, grupo_aula_default, by_code, by_name)
            # Si la actividad es de tipo especial y no hay aula explícita,
            # el código institucional del aula ES el tipo (LAB, INF…),
            # no el aula por defecto del grupo (que sería un aula teórica).
            # EXP y AD usan el aula del grupo cuando no tienen aula explícita.
            if not aula and tipo and tipo in ('LAB', 'INF', 'TLL', 'SEM'):
                classroom = tipo

            # Normalizar subgrupo: '' y 'todos' → vacío (equivale a toda la clase)
            subgrupo_norm = '' if subgrupo in ('', 'todos') else subgrupo

            key = (nombre, codigo, typology, dia, franja_label, subgrupo_norm, classroom)

            consolidated[key].append(real_week)

            if key not in key_meta:
                key_meta[key] = {
                    'nombre': nombre,
                    'codigo': codigo,
                    'typology': typology,
                    'dia': dia,
                    'franja_label': franja_label,
                    'subgrupo_norm': subgrupo_norm,
                    'classroom': classroom,
                    'observacion': observacion,
                    'curso': curso,
                    'grupo_num': grupo_num,
                    'subgrupos_numericos': subgrupos_numericos,
                }

        # ── Construir filas finales ──────────────────────────────────────────
        for key, real_weeks in consolidated.items():
            meta = key_meta[key]
            nombre = meta['nombre']
            codigo = meta['codigo']
            typology = meta['typology']
            dia = meta['dia']
            franja_label = meta['franja_label']
            subgrupo_norm = meta['subgrupo_norm']
            classroom = meta['classroom']
            observacion = meta['observacion']

            # Semanas únicas y ordenadas; se aplica WEEK_OFFSET para que
            # la semana interna 1 se exporte como 101, la 2 como 102, etc.
            weeks_sorted = sorted(set(real_weeks))
            weeks_str = ','.join(str(w + WEEK_OFFSET) for w in weeks_sorted)

            # HOURBEGIN / HOUREND
            hour_begin, hour_end = parse_franja(franja_label)

            # ID_PLAN: primeros 4 dígitos del código
            id_plan = int(codigo[:4]) if len(codigo) >= 4 and codigo[:4].isdigit() else codigo[:4]

            # SUBGRUPO y STUDENTGROUPSECTION
            if subgrupo_norm == '':
                subgrupo_col = '*'
                # Listar todos los subgrupos numéricos del grupo
                if meta['subgrupos_numericos']:
                    sections = ', '.join(
                        section_code(acronym, meta['curso'], meta['grupo_num'], sg)
                        for sg in meta['subgrupos_numericos']
                    )
                else:
                    # Si no hay subgrupos detectados, usar código genérico
                    sections = section_code(acronym, meta['curso'], meta['grupo_num'], '*')
            else:
                subgrupo_col = subgrupo_norm
                sections = section_code(acronym, meta['curso'], meta['grupo_num'], subgrupo_norm)

            all_rows.append({
                'EVENTNAME':            eventname(nombre),
                'EVENTTYPE':            'Clases',
                'DAYNUMBER':            DAY_NUMBER.get(dia, ''),
                'DAY':                  DAY_EN.get(dia, dia),
                'HOURBEGIN':            hour_begin,
                'HOUREND':              hour_end,
                'TYPOLOGYNAME':         typology,
                'ID_PLAN':              id_plan,
                'MODULECODE':           codigo,
                'SUBGRUPO':             subgrupo_col,
                'STUDENTGROUPSECTION':  sections,
                'TEACHERCODE':          '',
                'CLASSROOMS':           classroom,
                'WEEKS':                weeks_str,
                'ANNOTATIONS':          observacion,
            })

    # Ordenar filas: por código asignatura, día, franja, grupo
    all_rows.sort(key=lambda r: (
        r['MODULECODE'], r['TYPOLOGYNAME'],
        r['DAYNUMBER'], r['HOURBEGIN'],
        r['STUDENTGROUPSECTION'], r['SUBGRUPO']
    ))

    return all_rows


# ═══════════════════════════════════════════════════════════════════════════════
# ESCRITURA XLSX
# ═══════════════════════════════════════════════════════════════════════════════

def _make_border(style='thin'):
    s = Side(style=style, color='000000')
    return Border(left=s, right=s, top=s, bottom=s)


def write_xlsx(rows, output_path):
    """Escribe el Excel institucional en output_path."""
    if Workbook is None:
        raise ImportError("Se requiere openpyxl: pip install openpyxl --break-system-packages")

    wb = Workbook()
    ws = wb.active
    ws.title = 'Horarios'

    # Estilos de cabecera
    header_font = Font(bold=True, size=10, name='Arial', color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2E4099')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = _make_border('thin')

    # Anchos de columna aproximados
    col_widths = [30, 9, 11, 12, 10, 9, 13, 9, 14, 10, 45, 13, 18, 30, 20]

    # Escribir cabecera
    for col_idx, (header, width) in enumerate(zip(HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Escribir datos
    data_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')

    CENTER_COLS = {'EVENTTYPE', 'DAYNUMBER', 'DAY', 'HOURBEGIN', 'HOUREND',
                   'TYPOLOGYNAME', 'ID_PLAN', 'SUBGRUPO'}

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(HEADERS, start=1):
            value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9, name='Arial')
            cell.border = border
            cell.alignment = center_align if header in CENTER_COLS else data_align

    # Auto-filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}1'

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL (uso desde servidor)
# ═══════════════════════════════════════════════════════════════════════════════

def exportar(db_path, config_path, weeks_path, output_path,
             classrooms_json_path=None, tipos_json_path=None):
    """
    Genera el Excel institucional.

    Parámetros:
      db_path              Ruta al .db del grado
      config_path          Ruta al config.json (o directorio del grado)
      weeks_path           Ruta al weeks.xls o weeks.json de referencia.
                           Si se pasa un .xls y no existe el .json equivalente,
                           se genera automáticamente junto al .xls.
      output_path          Ruta destino del .xlsx generado
      classrooms_json_path Ruta a classrooms.json (autodetectado si None)
      tipos_json_path      Ruta a tipos_actividad.json (autodetectado si None)

    Devuelve output_path.
    """
    project_root = SCRIPT_DIR

    # Autodetectar ficheros auxiliares si no se pasan
    if classrooms_json_path is None:
        classrooms_json_path = os.path.join(project_root, 'config', 'classrooms.json')
    if tipos_json_path is None:
        tipos_json_path = os.path.join(project_root, 'config', 'tipos_actividad.json')

    # Cargar recursos
    weeks_map = load_weeks_map(weeks_path)       # genera weeks.json si hace falta
    config = load_config(config_path)
    by_code, by_name = load_classrooms(classrooms_json_path)
    af_to_code = load_tipos_actividad(tipos_json_path)

    # Consultar DB
    grupos, data_by_grupo = fetch_data(db_path)

    # Generar filas
    rows = build_rows(grupos, data_by_grupo, config, weeks_map, by_code, by_name, af_to_code)

    # Escribir Excel
    write_xlsx(rows, output_path)

    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# EJECUCIÓN STANDALONE
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Exporta horarios Janux al formato Excel institucional UPCT.'
    )
    parser.add_argument(
        '--grado', required=True,
        help='Ruta al directorio del grado (contiene config.json y el .db)'
    )
    parser.add_argument(
        '--weeks', default=None,
        help='Ruta a weeks.xls o weeks.json (por defecto: weeks.xls en la raíz del proyecto). '
             'Si se pasa un .xls y no existe el .json equivalente, se genera automáticamente.'
    )
    parser.add_argument(
        '--output', default=None,
        help='Ruta del xlsx de salida (por defecto: exportacion_institucional.xlsx en el grado)'
    )
    parser.add_argument(
        '--classrooms', default=None,
        help='Ruta a classrooms.json (por defecto: config/classrooms.json)'
    )
    parser.add_argument(
        '--tipos', default=None,
        help='Ruta a tipos_actividad.json (por defecto: config/tipos_actividad.json)'
    )
    args = parser.parse_args()

    # Resolver rutas
    grado_dir = args.grado
    config_file = os.path.join(grado_dir, 'config.json')

    if not os.path.exists(config_file):
        print(f"ERROR: No se encuentra config.json en {grado_dir}")
        sys.exit(1)

    # Buscar DB
    with open(config_file, encoding='utf-8') as f:
        cfg = json.load(f)
    db_name = cfg.get('server', {}).get('db_name', 'horarios.db')
    db_path = os.path.join(grado_dir, db_name)
    if not os.path.exists(db_path):
        dbs = [f for f in os.listdir(grado_dir) if f.endswith('.db')]
        if dbs:
            db_path = os.path.join(grado_dir, dbs[0])
        else:
            print(f"ERROR: No se encuentra ningún .db en {grado_dir}")
            sys.exit(1)

    # Localizar weeks (acepta .xls o .json)
    # Orden de búsqueda por defecto: config/weeks.json → weeks.xls (raíz)
    weeks_path = args.weeks
    if not weeks_path:
        json_candidate = os.path.join(SCRIPT_DIR, 'config', 'weeks.json')
        xls_candidate  = os.path.join(SCRIPT_DIR, 'weeks.xls')
        weeks_path = json_candidate if os.path.exists(json_candidate) else xls_candidate
    if not os.path.exists(weeks_path):
        print(f"ERROR: No se encuentra el fichero de semanas: {weeks_path}")
        print("Especifica la ruta con --weeks (weeks.xls o config/weeks.json)")
        sys.exit(1)

    output_path = args.output or os.path.join(grado_dir, 'exportacion_institucional.xlsx')
    classrooms_path = args.classrooms or os.path.join(SCRIPT_DIR, 'config', 'classrooms.json')
    tipos_path = args.tipos or os.path.join(SCRIPT_DIR, 'config', 'tipos_actividad.json')

    print(f"  Grado:      {grado_dir}")
    print(f"  BD:         {db_path}")
    print(f"  Semanas:    {weeks_path}")
    print(f"  Salida:     {output_path}")
    print("Generando...")

    try:
        exportar(db_path, config_file, weeks_path, output_path,
                 classrooms_path, tipos_path)
        print(f"✓ Exportación completada: {output_path}")
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
